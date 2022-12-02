import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from utils import *


def create_master_xls(matched_hint_files, export_path, hints_path, bot_hints_path):
    """
    Takes the list of matched hint file paths and creates a master xls workbook, with named sub-sheets for each match.
    @param [list] matched_hint_files: the list of hint file paths that are used in the TA.
    @param [str] export_path: the path to the export folder, used when saving the master xls and also in the optimize
    file name function.
    @param [str] hints_path: the path to the hints folder, used in the optimize file name function.
    @param [str] bot_hints_path: the path to the bot_hints folder, used in the optimize file name function.
    @return [obj] master_xls_obj: the master xls as an Excel object.
    """
    master_xls_obj = Workbook()
    overview_sheet = master_xls_obj.active
    overview_sheet.title = 'Overview'
    print('Created master spreadsheet.')

    for file in matched_hint_files:
        target_sheet_name = optimize_file_name(file, bot_hints_path, hints_path)
        master_xls_obj.create_sheet(title=target_sheet_name)
        print(f'Created Sheet in Master: {target_sheet_name}')

    master_xls_obj.save(filename=export_path + '/master.xlsx')

    return master_xls_obj


def get_csv_paths(matched_hint_files, hints_path):
    """
    Takes the list of matched hint file names and returns their full csv paths.
    @param [list] matched_hint_files: the list of hint file paths that are used in the TA.
    @param [str] hints_path: the path to the hints folder.
    @return [list] csv_file_paths: a list of csv file paths that are read into for data, and reworked into xlsx paths.
    """
    csv_file_paths = []
    for csv_file_name in matched_hint_files:
        csv_path = hints_path + '/' + csv_file_name
        csv_file_paths.append(csv_path)

    return csv_file_paths


def create_target_xls_paths(csv_file_paths):
    """
    Takes the list of csv file paths and reworks the path strings into xlsx file paths.
    @param [list] csv_file_paths: the list of matched hint files as csv paths.
    @return [list] xl_file_paths: the list of matched hint files as xlsx paths.
    """
    xl_file_paths = []
    for path in csv_file_paths:
        change_to_xlsx_path = path[:-4] + '.xlsx'
        final_xl_file_path = change_to_xlsx_path.replace('hints', 'bot-hints')
        xl_file_paths.append(final_xl_file_path)

    return xl_file_paths


def create_xls_from_csv_paths(csv_file_paths, xl_file_paths):
    """
    Reads into the csv file paths, extracts all data from them, and transfers the data to a newly created xlsx file.
    @param [list] csv_file_paths: the list of matched hint files as csv paths.
    @param [list] xl_file_paths: the list of matched hint files as xlsx paths.
    """
    for index, csv_file_path in enumerate(csv_file_paths):
        csv_data = []
        with open(csv_file_path) as file_obj:
            reader = csv.reader(file_obj)
            for row in reader:
                csv_data.append(row)
        workbook = Workbook()
        sheet = workbook.active
        for row in csv_data:
            sheet.append(row)
        print(f'Created Excel file: {xl_file_paths[index]}')
        workbook.save(xl_file_paths[index])


def get_all_xls_data(xls_paths, bot_hints_path, hints_path):
    """
    Extracts all data from every newly created xlsx. Puts that data as a list into a dictionary that uses the file name
    as the key. ({file_name: [sheet_data]})
    @param [list] xls_paths: the list of matched hint files as xlsx paths.
    @param [str] bot_hints_path: the path to the bot_hints folder.
    @param [str] hints_path: the path to the hints folder.
    @return [dict] xls_data: A dict with a list of all data for every xlsx, with the
    file name as the dict key. {file_name: [sheet_data]}
    """
    xls_data = []
    for path in xls_paths:
        target_name = optimize_file_name(path, bot_hints_path, hints_path)
        file = load_workbook(filename=path)
        sheet = file.active
        sheet_dict, sheet_data = {}, []
        for row in sheet.rows:
            sheet_data.append(row)
        sheet_dict.update({target_name: sheet_data})
        xls_data.append(sheet_dict)

    return xls_data


def transfer_xls_data_to_master(xls_data, master_xls_obj, export_path):
    """
    Takes the xls data dict and transfers it to it's corresponding sheet in the master.xlsx.
    @param [dict] xls_data: A dict with a list of all data for every matched xlsx, with the
    file name as the dict key. {target_name: [sheet_data]}
    @param [obj] master_xls_obj: the master xls as an Excel object.
    @param [str] export_path: the path to the export folder.
    """
    for entry in xls_data:
        data_key = list(entry.keys())[0]
        entry_data = list(entry.values())[0]

        print(f'Added data to {data_key} in master spreadsheet.')

        for i in entry_data:  # Accesses each element of tuple (columns of xl object)
            data_values = [data.value for data in i]  # Makes list of values instead of cell objects
            for sheet in master_xls_obj:
                if data_key == sheet.title:
                    target_sheet = master_xls_obj[data_key]  # If key and sheet title match, appends data values.
                    target_sheet.append(data_values)

    master_xls_obj.save(filename=export_path + '/master.xlsx')


def get_master_xls_obj(master_xls_path):
    """
    Returns an existing master xls object if one has already been created.
    @param [str] master_xls_path: path to the master xls file.
    @return [obj] master_xls_obj: the master xls as an Excel object.
    """
    master_xls_obj = load_workbook(master_xls_path)

    return master_xls_obj

